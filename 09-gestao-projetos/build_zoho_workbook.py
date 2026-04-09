"""Generate XLSX for Zoho Projects import + director integration backlog."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "porto-coast-stays-zoho-projects.xlsx"

# Zoho Projects task import: common CSV columns (align CSV export from this sheet)
ZOHO_HEADERS = [
    "Task List Name",
    "Task Name",
    "Description",
    "Priority",
    "Start Date",
    "End Date",
    "Milestone",
    "Tags",
    "Owner Email",
    "Duration (days)",
    "Dependency (Task Name)",
]

# Priority: Zoho accepts None, Low, Medium, High — map from business rank
def row(tasklist, name, desc, prio, start="", end="", milestone="", tags="", owner="", duration="", dep=""):
    return [
        tasklist,
        name,
        desc,
        prio,
        start,
        end,
        milestone,
        tags,
        owner,
        duration,
        dep,
    ]

TASKS = []

# --- Per-property onboarding (template: duplicate project per unit) ---
for bloco, tasks in [
    (
        "B1 Preparação física",
        [
            ("Visita e aprovação do imóvel", "Estado, equipamentos, acesso prédio, RNAL se aplicável", "High"),
            ("Instalar Nuki + Keypad + Bridge", "Bloqueia automação de códigos e check-in", "High"),
            ("Configurar acesso porta do prédio", "Teclado ou integração documentada", "High"),
            ("Auditoria WiFi e climatização", "Mín. 50 Mbps; AC/aquecimento testado", "Medium"),
            ("Stock cozinha e WC", "Café, detergentes, papel, amenities", "Medium"),
            ("Roupa de cama e toalhas (2–3 sets)", "Qualidade hotel", "Medium"),
            ("Kit boas-vindas", "Água, snack local, café cápsula", "Low"),
            ("Limpeza pré-fotos", "Imóvel 100% pronto para sessão", "High"),
        ],
    ),
    (
        "B2 Fotos profissionais",
        [
            ("Contratar fotógrafo AL", "Orçamento 150–300 €; briefing ordem das fotos", "High"),
            ("Preparar set do dia de fotos", "Luz natural, sem clutter, styling", "High"),
            ("Sessão e seleção 15–25 fotos", "Horizontal; ordem capa→sala→quarto→cozinha→diferenciador", "High"),
            ("Entrega final validada", "Consistência visual; aprovação gestor", "High"),
        ],
    ),
    (
        "B3 Listing e OTAs",
        [
            ("Listing Airbnb (título, hook, comodidades)", "Instant Book; calendário 4–6 meses; pricing lançamento -20/-30%", "High"),
            ("Listing Booking.com", "Replicar conteúdo; sync via PMS", "High"),
            ("Regras da casa e políticas", "Silêncio, não fumar, máx. hóspedes", "Medium"),
            ("Descontos semanal/mensal", "5–10% / 15–20%; estadia mín. 1–2 noites", "Medium"),
        ],
    ),
    (
        "B4 Sistemas e automações",
        [
            ("PMS + Channel Manager", "Hostaway/Smoobu; Airbnb+Booking ligados", "High"),
            ("Integração Nuki ↔ PMS", "Códigos por reserva; janelas check-in/out", "High"),
            ("Sequência mensagens automáticas (9 triggers)", "Ver roadmap / fluxo-completo", "High"),
            ("Notificação limpeza pós-checkout", "PMS → WhatsApp operações", "High"),
            ("Pricelabs ligado ao PMS", "Base, min/max, fds +15–25%, eventos +30–80%", "Medium"),
            ("Guia digital do apartamento", "Notion/PDF/Hostaway; link nas mensagens", "Medium"),
            ("Faturação InvoiceXpress/Moloni + e-fatura", "IVA 6%; integração API PMS", "High"),
            ("SIBA + registo hóspedes", "Portal ou Chekin; teste end-to-end", "High"),
            ("Relatório mensal proprietário", "Export PMS; envio até dia 5", "Medium"),
        ],
    ),
    (
        "B5 Go-live",
        [
            ("Reserva de teste end-to-end", "Código, mensagens, calendário, limpeza notificada", "High"),
            ("Validar anti double-booking", "Airbnb vs Booking vs PMS", "High"),
            ("Limpeza + fallback ativos", "Contrato backup <30 min", "High"),
            ("Publicar listing e monitorizar 72h", "Resposta <1h; aceitação >90%", "High"),
        ],
    ),
    (
        "B6 Primeiras reviews (0–30d)",
        [
            ("SLA mensagens <1h", "Notificações push PMS + Airbnb", "High"),
            ("Pedido review D+1 (automático)", "Template simples pós-checkout", "Medium"),
            ("Responder reviews <48h", "Todas as plataformas", "Medium"),
            ("Ajuste pricing pós 3–5 reviews", "+10%; depois Pricelabs pleno", "Medium"),
            ("Meta: 10 reviews >4.8 e ocupação >65%", "Rever semanalmente KPIs", "High"),
        ],
    ),
]:
    dep = ""
    for tname, tdesc, prio in tasks:
        TASKS.append(row(bloco, tname, tdesc, prio, milestone=bloco, tags="onboarding;AL;PortoCoast"))
        dep = tname  # simple chain hint for same list (optional)

# --- Roadmap phases (company-level project) ---
PHASE_TASKS = [
    row("Fase 1 Lançamento (0–30d)", "Fechar Blocos 1–6 imóvel 1", "Marco: primeira reserva", "High", milestone="Fase 1"),
    row("Fase 1 Lançamento (0–30d)", "Replicar onboarding imóvel 2", "Mesmo template de projeto", "High", milestone="Fase 1"),
    row("Fase 2 Validação (30–90d)", "Pricelabs pleno + análise ocupação/ADR", "Dashboard semanal", "Medium", milestone="Fase 2"),
    row("Fase 2 Validação (30–90d)", "Documentar SOPs para delegação", "Limpeza, suporte, onboarding", "Medium", milestone="Fase 2"),
    row("Fase 3 Expansão (3–6m)", "Pipeline proprietários + proposta com dados", "10–15 abordagens", "Medium", milestone="Fase 3"),
    row("Fase 3 Expansão (3–6m)", "Contratar suporte part-time", "Inbox e escalations", "Low", milestone="Fase 3"),
    row("Fase 4 Sistema (6–12m)", "Operations manager + limpeza dedicada", "10+ unidades", "Low", milestone="Fase 4"),
]
TASKS.extend(PHASE_TASKS)

# --- Director: integration & tooling programme ---
INTEGRATION = [
    row(
        "Direção — Integrações",
        "Zoho Projects: projeto-mãe + template por imóvel",
        "1 projeto 'Porto Coast Stays — Operações'; duplicar por unidade com datas.",
        "High",
        milestone="Stack direção",
        tags="zoho;governance",
    ),
    row(
        "Direção — Integrações",
        "Zoho CRM ↔ Projetos",
        "Deal/Account (proprietário) → Projeto (onboarding) via Flow/Deluge ou Zoho Flow.",
        "High",
        milestone="Stack direção",
        tags="zoho;crm",
    ),
    row(
        "Direção — Integrações",
        "Zoho Books / Invoice + AL",
        "Separar faturação hóspede (InvoiceXpress/PMS) da contabilidade gestora (Books) se aplicável.",
        "Medium",
        milestone="Stack direção",
        tags="faturação;books",
    ),
    row(
        "Direção — Integrações",
        "GitHub: repo docs + CI leve",
        "PR checklist para alterações a SOPs; branch protection em main.",
        "Medium",
        milestone="Stack direção",
        tags="github;docs",
    ),
    row(
        "Direção — Integrações",
        "Netlify / Vercel: deploy site marca",
        "Ligar repo → preview PR → produção; secrets em painel, não no repo.",
        "Medium",
        milestone="Stack direção",
        tags="netlify;vercel",
    ),
    row(
        "Direção — Integrações",
        "Supabase: apenas se produto precisar DB",
        "Reservas ficam no PMS; Supabase para apps próprios (ex. guest app).",
        "Low",
        milestone="Stack direção",
        tags="supabase",
    ),
    row(
        "Direção — Integrações",
        "Painel semanal direção (read-only)",
        "KPIs: resposta, ocupação, reviews, limpezas OK — agregador (Sheets ou Analytics).",
        "High",
        milestone="Stack direção",
        tags="kpi;dashboard",
    ),
]
TASKS.extend(INTEGRATION)


def style_header(ws, row_idx=1):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, max_width=55):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for r in range(1, min(ws.max_row + 1, 200)):
            v = ws.cell(row=r, column=col).value
            if v:
                maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[letter].width = min(max_width, max(12, maxlen + 2))


def main():
    wb = Workbook()

    # Sheet 1: Import-ready
    ws1 = wb.active
    ws1.title = "Zoho_Import_Tasks"
    ws1.append(ZOHO_HEADERS)
    for t in TASKS:
        ws1.append(t)
    style_header(ws1)
    autosize(ws1)

    # Sheet 2: Project shell for Zoho
    ws2 = wb.create_sheet("Zoho_Project_Setup")
    ws2.append(["Field", "Value / guidance"])
    rows_setup = [
        ("Suggested Project Name", "Porto Coast Stays — Onboarding Imóvel (template)"),
        ("Project Type", "Template — duplicate per property"),
        ("Default Task layout", "Task Lists = B1…B6 + Fases + Direção"),
        ("Import path", "Zoho Projects → Tasks → Import (CSV export from Zoho_Import_Tasks)"),
        ("API next step", "OAuth Zoho Projects + store refresh token in vault (not repo)"),
        ("Owner model", "Director = project owner; executors = task assignees only"),
    ]
    for a, b in rows_setup:
        ws2.append([a, b])
    style_header(ws2)
    autosize(ws2)

    # Sheet 3: Integration architecture (director view)
    ws3 = wb.create_sheet("Arquitetura_Integracoes")
    ws3.append(["Camada", "Ferramenta", "Papel", "Sincronização recomendada"])
    arch = [
        ("Comercial", "Zoho CRM", "Leads proprietários, contas, pipeline gestão", "Zoho Flow → cria projeto em Projects ao ganhar deal"),
        ("Execução", "Zoho Projects", "Onboarding imóvel, SOP tasks, marcos", "API ou import CSV semanal; status → CRM custom field"),
        ("Operação hóspede", "Hostaway + OTAs", "Reservas, mensagens, limpeza", "Não duplicar no CRM; só alertas exceção"),
        ("Faturação AL", "InvoiceXpress / e-fatura", "Fatura por reserva, AT", "Webhook/API PMS; Books só consolidado mensal"),
        ("Código / Site", "GitHub", "Website, documentação", "Actions → Netlify/Vercel; PR = revisão diretor"),
        ("Infra site", "Netlify / Vercel", "Deploy estático", "Branch main = prod; preview = PR"),
        ("Dados produto", "Supabase", "Só se houver app próprio", "MCP/Edge; não substitui PMS"),
    ]
    for r in arch:
        ws3.append(list(r))
    style_header(ws3)
    autosize(ws3)

    # Sheet 4: CSV export instructions
    ws4 = wb.create_sheet("Instrucoes_Import")
    instr = [
        "1. No Excel: folha 'Zoho_Import_Tasks' → Guardar como CSV (UTF-8) se o Zoho pedir ficheiro.",
        "2. Em Zoho Projects: criar projeto → Task Lists com os mesmos nomes que 'Task List Name' (ou mapear na importação).",
        "3. Prioridades no ficheiro: High / Medium / Low (ajustar ao teu dicionário Zoho se diferente).",
        "4. Datas: preencher Start/End quando planificares; vazio = tarefas backlog.",
        "5. Owner Email: preencher emails Zoho dos executores; diretor fica com visão por projeto/marco.",
        "6. Para API: usar Zoho Projects REST v3 — autenticação OAuth2; não commits de tokens no GitHub.",
        "7. Direção: automatizar criação de projeto a partir de CRM (Zoho Flow) + checklist este XLSX como template.",
    ]
    ws4.append(["Instrução"])
    for line in instr:
        ws4.append([line])
    style_header(ws4)
    autosize(ws4)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
